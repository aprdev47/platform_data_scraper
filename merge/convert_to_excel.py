#!/usr/bin/env python3
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Read CSV file
csv_file = '/Users/athul/dev/ai exp/merge/hr_payroll_platforms_contacts.csv'
excel_file = '/Users/athul/dev/ai exp/merge/hr_payroll_platforms.xlsx'

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HR Payroll Platforms"

# Read CSV and write to Excel
with open(csv_file, 'r', encoding='utf-8') as f:
    csv_reader = csv.reader(f)
    for row_idx, row in enumerate(csv_reader, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Format header row
            if row_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

# Adjust column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 45

# Freeze the header row
ws.freeze_panes = "A2"

# Save the workbook
wb.save(excel_file)
print(f"Excel file created successfully: {excel_file}")
